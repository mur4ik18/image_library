from openpyxl import Workbook, drawing, worksheet
import zipfile
from glob import glob
import os
from tutorial.settings import *

class Excelt:
    def __init__(self, name_file):
        self.fileName: str = name_file
        self.wb = Workbook()
        self.ws = None

    def create_headers(self, headers=None):
        self.ws = self.wb.active
        self.ws.append(headers)
        self.ws.freeze_panes = "A1"
        self.ws.auto_filter.ref = "A1:O100"

    # methods
    def write_data(self, data, row):
        self.ws.cell(column=1, row=row, value=str(data['numbers']))
        self.ws.cell(column=2, row=row, value=data['account'])
        self.ws.cell(column=3, row=row, value=data['post_url'])
        img = drawing.image.Image(data['image'])
        img.height = img.height / 6
        img.width = img.width / 6
        img.anchor = f'D{row}'
        self.ws.column_dimensions['E'].width = img.width / 7
        rd = self.ws.row_dimensions[row]  # get dimension for row 3
        rd.height = img.height - 40
        self.ws.add_image(img)
        self.ws.cell(column=5, row=row, value=data['date'])
        self.ws.cell(column=6, row=row, value=data['likes'])
        self.ws.cell(column=7, row=row, value=data['comments'])
        self.ws.cell(column=8, row=row, value=data['caption'])
        self.ws.cell(column=9, row=row, value=data['hashtags'])
        self.ws.cell(column=10, row=row, value=data['mentions'])
        self.ws.cell(column=11, row=row, value=data['text_comments'])
        self.ws.cell(column=12, row=row, value=data['Image_credit'])
        self.ws.cell(column=13, row=row, value=data['tags'])
        self.ws.cell(column=14, row=row, value=f"https://www.instagram.com/{data['account']}/")
        address = os.environ.get("ADDRESS",'localhost')
        self.ws.cell(column=15, row=row, value=f"http://34.208.90.101/images/{data['image']}")

    def save_file(self):
        try:
            os.mkdir('files')
        except FileExistsError:
            print('skip')
        self.wb.save(f'files/{self.fileName}.xlsx')

def download_data(file_name, data):
    header_names = ['numbers', 'account', 'post_url', 'image', 'date', 'likes', 'comments',
                            'caption', 'hashtags', 'mentions', 'text_comments', 'Image_credit','tags','instagram_account','image_link']
    Excel = Excelt(file_name)

    Excel.create_headers(headers=header_names)
    t = 2     
    
    # with zipfile.ZipFile(f'files/{file_name}.zip', 'w', zipfile.ZIP_DEFLATED) as myzip:
        
    for i in data:
        Excel.write_data(data=data[i], row=t)
        # shutil.copyfile(f"{MEDIA_URL}/{data[i]['image']}", f'files/images/{t-1}.jpg')
        # myzip.write(f'files/images/{t-1}.jpg')
        t+=1
        
    Excel.save_file() 

        # myzip.write(f"files/{file_name}.xlsx")
    